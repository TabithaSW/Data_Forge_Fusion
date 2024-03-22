
"""

Py file to test json, xml, csv, parquet, excel convert.
PDF soon?

"""
# basic data manipulation libraries
import pandas as pd
import numpy as np
import teradatasql

import matplotlib.pyplot as plt
import seaborn as sns

# File conversion libraries:
import xml.etree.ElementTree as ET
import xml.dom.minidom

import json
import csv
import openpyxl
import operator
import os
import dask.dataframe as dd

#GUI Stuff
# GUI library:
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk
import tkinter.scrolledtext as tkst
import tkinterdnd2

# NEW FORMAT, EXCEL:
def write_excel_file(datalist,filepath=None):
    # allow users to choose filename
    if not filepath:
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel File", ".xlsx")])

    # create excel template workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    # headers 
    headers = list(datalist[0].keys())
    sheet.append(headers)

    for dict_ in datalist:
        sheet.append([dict_[header] for header in headers])
    workbook.save(filepath)

# changed to using openpyxl because of pandas limit
def read_excel_file(file_path):
    # excel files are workbooks compirsed of individual sheets of data
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]

    datalist = [] #all my read funcs convert to a python list of dictionaries.
    # start at second row, pass header.
    for row in sheet.iter_rows(min_row = 2, values_only = True):
        row_dict = dict(zip(headers,row)) # need to zip headers if multiple sheets
        datalist.append(row_dict)
    return datalist


# Reads in a CSV File
def read_csv_file(filename):
    # Dictionary Object to store data:
    data = []
    # Open CSV File:
    with open(filename, 'r') as file:
        # Pull data from CSV:
        reader = csv.DictReader(file)
        # Loop through file records,
        for row in reader:
            data.append(row)
    return data

# Writes to a CSV File Format
def write_csv_file(filename, data):
    """
    Takes a filename (to be writen to) and a data object 
    (created by one of the read_*_file functions). 
    Writes the data in the CSV format.
    """
    filename = None
    if not filename:
        filename = filedialog.asksaveasfilename(defaultextension=".csv",filetypes=[("CSV File",".csv")])
    # Generate instance of new file to write:
    with open(filename, "w", newline="") as outfile:
        # Pass CSV file and dictionary keys as column headers:
        writer = csv.DictWriter(outfile, fieldnames=data[0].keys())

        writer.writeheader()

        for entry in data:
            writer.writerow(entry)

# Reads in a JSON File
def read_json_file(filename):
    """
    Similar to read_csv_file, except works for JSON files.
    """
    with open(filename) as json_file:
        data = json.load(json_file)
    return [data]

# write json new 2024
def write_json_file(data, filename=None):
    """
    Writes JSON files with formatted indentation.
    Assumes data is a list and writes the first element to the file.
    """
    if not filename:
        filename = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON File", ".json")])
    
    with open(filename, 'w') as file:
        if isinstance(data, list) and len(data) > 0:
            json.dump(data[0], file, indent=4)  # Writes only the first dictionary element of the list
        else:
            json.dump(data, file, indent=4)  # Handle non-list data normally

# Reads an XML file
def read_xml_file(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()
    data_list = []
    for record in root:
        record_dict = {}
        for column in record:
            record_dict.update({column.tag: column.text})
        data_list.append(record_dict)
    return data_list

def write_xml_file(data_list, filename=None):
    if not filename:
        filename = filedialog.asksaveasfilename(defaultextension=".xml", filetypes=[("XML File", ".xml")])

    root = ET.Element("data")
    for record in data_list:
        record_node = ET.SubElement(root, "record")
        for column, value in record.items():
            column_node = ET.SubElement(record_node, column)
            column_node.text = value

    tree = ET.ElementTree(root)
    # Convert the ElementTree to a string and then parse it with minidom for pretty printing
    xml_string = ET.tostring(root, encoding='utf-8')
    parsed_xml = xml.dom.minidom.parseString(xml_string)
    pretty_xml_as_string = parsed_xml.toprettyxml(indent="  ")

    with open(filename, 'w', encoding='utf-8') as file:
        file.write(pretty_xml_as_string)



# new file type, parquet
# Parquet is one of the fastest file types to read generally and much faster than either JSON or CSV.
def read_parquet_file(filename):
    """
    Reads data from a Parquet file into a list of dictionaries (similar to other read functions).
    """
    df = pd.read_parquet(filename)
    return df.to_dict('records')

def write_parquet_file(data_list, filename=None):
    """
    Writes a list of dictionaries to a Parquet file, allowing users to choose the filename.
    """
    if filename is None:
        filename = filedialog.asksaveasfilename(defaultextension=".parquet", filetypes=[("Parquet File", ".parquet")])
    
    df = pd.DataFrame(data_list)
    df.to_parquet(filename, index=False)


if __name__ == "__main__":
    
    """
    Test Parquet
    """
    # test read
    #test_p = read_parquet_file('Test_Folder/sample_data.parquet')
    #print(test_p)

    # test write 
    #test_wp = write_parquet_file(test_p)

    """
    Test CSV
    """
    #csv_test = read_csv_file('Test_Folder/diamonds.csv')
    #print(csv_test[0])

    #csv_w = write_csv_file(data=csv_test)

    """
    Test XML
    """
    #XML_TEST = read_xml_file('Test_Folder/sample_xml.xml')
    #print(XML_TEST)

    #xml_w = write_xml_file(XML_TEST)

    """
    Test JSON
    """

    #json_test = read_json_file('Test_Folder/join_test.json')
    #print(json_test)

    #j_w = write_json_file(json_test)

    """
    Test Excel
    """

    #excel_test = read_excel_file('Test_Folder/fruit_excel.xlsx')
    #print(excel_test)

    #ex_w = write_excel_file(excel_test)