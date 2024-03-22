
# basic data manipulation libraries
import pandas as pd
import numpy as np
import teradatasql

import matplotlib.pyplot as plt
import seaborn as sns

# File conversion libraries:
import xml.etree.ElementTree as ET
import xml.dom.minidom

import json
import csv
import openpyxl
import operator
import os
import dask.dataframe as dd

#GUI Stuff
# GUI library:
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk
import tkinter.scrolledtext as tkst
import tkinterdnd2

# Need some small funcs from main py
import DataConverterMain

# Reads in a CSV File
def read_csv_file(filename):
    # Dictionary Object to store data:
    data = []
    # Open CSV File:
    with open(filename, 'r') as file:
        # Pull data from CSV:
        reader = csv.DictReader(file)
        # Loop through file records,
        for row in reader:
            data.append(row)
    return data

# Writes to a CSV File Format
def write_csv_file(data,filename=None):
    """
    Takes a filename (to be writen to) and a data object 
    (created by one of the read_*_file functions). 
    Writes the data in the CSV format.
    """
    #filename = None
    if not filename:
        filename = filedialog.asksaveasfilename(defaultextension=".csv",filetypes=[("CSV File",".csv")])
    # Generate instance of new file to write:
    with open(filename, "w", newline="") as outfile:
        # Pass CSV file and dictionary keys as column headers:
        writer = csv.DictWriter(outfile, fieldnames=data[0].keys())

        writer.writeheader()

        for entry in data:
            writer.writerow(entry)

# Reads in a JSON File
def read_json_file(filename):
    """
    Similar to read_csv_file, except works for JSON files.
    """
    with open(filename) as json_file:
        data = json.load(json_file)
    return [data]

# Writes to a JSON File Format
def write_json_file(data, filename=None):
    """
    Writes JSON files with formatted indentation.
    """
    if not filename:
        filename = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON File", ".json")])
    
    with open(filename, 'w') as file:
        json.dump(data, file, indent=4)  # Writes the entire data list to the file

# Reads an XML file
def read_xml_file(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()
    data_list = []
    for record in root:
        record_dict = {}
        for column in record:
            record_dict.update({column.tag: column.text})
        data_list.append(record_dict)
    return data_list

# Writes to an XML File:
def write_xml_file(data_list, filename=None):
    if not filename:
        filename = filedialog.asksaveasfilename(defaultextension=".xml", filetypes=[("XML File", ".xml")])

    root = ET.Element("data")
    for record in data_list:
        record_node = ET.SubElement(root, "record")
        for column, value in record.items():
            column_node = ET.SubElement(record_node, column)
            column_node.text = value

    tree = ET.ElementTree(root)
    # Convert the ElementTree to a string and then parse it with minidom for pretty printing
    xml_string = ET.tostring(root, encoding='utf-8')
    parsed_xml = xml.dom.minidom.parseString(xml_string)
    pretty_xml_as_string = parsed_xml.toprettyxml(indent="  ")

    with open(filename, 'w', encoding='utf-8') as file:
        file.write(pretty_xml_as_string)

def detect_file(file_path):
        """
        Calls all the custom read functions for xml, csv, excel, parquet, json.
        
        """
        #print("DETECT FILE PATH TEST:",file_path)
        # When user is prompted to choose a file, we need to ensure it is correctly identified before conversion.
        x, file_extension = os.path.splitext(file_path) # X is root, rest is the extension, x is throwaway var.

        #print(file_extension,"FILE EX TEST")

        if file_extension == ".csv":
            return read_csv_file(file_path)
        elif file_extension == ".xml":
            return read_xml_file(file_path)
        elif file_extension == ".xlsx":
            return read_excel_file(file_path)
        elif file_extension == ".parquet":
            return read_parquet_file(file_path)
        else:
            return read_json_file(file_path)
        
def prompt_file_choice(file_path):
        file_options = ["CSV", "JSON", "XML","csv","json","xml","excel","EXCEL","PARQUET","parquet"]
        file_name = os.path.basename(file_path)
        user_choice = simpledialog.askstring(f"{file_name} Conversion Options","CSV, JSON, EXCEL, PARQUET, or XML?",initialvalue=file_options[0])
        if user_choice.lower() in [option.lower() for option in file_options]:
            return user_choice
        else:
            messagebox.showwarning("Conversion Cancelled")


# fixed 2024 excel
def write_excel_file(datalist,filepath=None):
    # allow users to choose filename
    if not filepath:
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel File", ".xlsx")])

    # create excel template workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    # headers 
    headers = list(datalist[0].keys())
    sheet.append(headers)

    for dict_ in datalist:
        sheet.append([dict_[header] for header in headers])
    workbook.save(filepath)


# changed to using openpyxl because of pandas limit
def read_excel_file(file_path):
    # excel files are workbooks comprised of individual sheets of data
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]

    datalist = [] #all my read funcs convert to a python list of dictionaries.
    # start at second row, pass header.
    for row in sheet.iter_rows(min_row = 2, values_only = True):
        row_dict = dict(zip(headers,row)) # need to zip headers if multiple sheets
        datalist.append(row_dict)
    return datalist

# new file type, parquet
# Parquet is one of the fastest file types to read generally and much faster than either JSON or CSV.
def read_parquet_file(filename):
    """
    Reads data from a Parquet file into a list of dictionaries (similar to other read functions).
    """
    df = pd.read_parquet(filename)
    return df.to_dict('records')

def write_parquet_file(data, filename=None):
    """
    Writes a list of dictionaries to a Parquet file, allowing users to choose the filename.
    """
    if not filename:
        filename = filedialog.asksaveasfilename(defaultextension=".parquet", filetypes=[("Parquet File", ".parquet")])
    
    df = pd.DataFrame(data)
    df.to_parquet(filename, index=False)


# Lets do some DBMS stuff that made go hand in hand with file conversion.
# Teradata to start:
def convert_teradata(server, username, password, query,output_path = None):
    #allow user to enter query
    try:
        # assign connection object
        with teradatasql.connect(host = server, user=username,password=password) as conn:
            # cursor to execute queries
            with conn.cursor() as cur:
                cur.execute(query)
                cols = [col[0] for col in cur.description]
                result = cur.fetchall() # store query results

        display_teradata_preview(query_result=result,headers=cols)

        # lets convert this file format first,
        csv_filename = filedialog.asksaveasfilename(defaultextension=".csv",filetypes=[("CSV File",".csv")])
        if output_path:
            csv_filename = os.path.join(output_path, csv_filename)
        with open(csv_filename,'w',newline='') as f:
            csv_write = csv.writer(f)
            # if header:
            if cur.description:
                csv_write.writerow([col[0] for col in cur.description])
            # complete write 
            csv_write.writerows(result)
            
        return csv_filename
    # if user enters illegible query, produce error by teradata.
    except teradatasql.Error as error_:
        return f"Error: {error_}"

# Teradata preview:
def display_teradata_preview(query_result,headers=None):
    # preview window
    df_prev = pd.DataFrame(query_result)

    prev_window = tk.Toplevel()
    prev_window.title("Result Preview")
    prev_window.geometry("1500x500")

    frame = ttk.Frame(prev_window)
    frame.pack(expand=True,fill="both")

    tree = ttk.Treeview(frame, show ="headings")
    tree["columns"] = list(df_prev.columns)

    # Initialize sort order for each column
    sort_orders = {col: "ascending" for col in df_prev.columns}

    def toggle_sort(column):
        nonlocal sort_orders  # Access the sort_orders variable in the outer scope

        # Determine the next sort order
        next_order = "descending" if sort_orders[column] == "ascending" else "ascending"

        # Sort the DataFrame
        sorted_df = df_prev.sort_values(by=column, ascending=(next_order == "ascending"))

        # Clear existing rows and insert sorted data
        for row in tree.get_children():
            tree.delete(row)
        for index, row in sorted_df.iterrows():
            tree.insert("", "end", values=list(row))
        
        # Toggle sort order for next click
        sort_orders[column] = next_order

    # Set up columns and headings with sorting functionality
    for col in df_prev.columns:
        tree.column(col, anchor="w")
        tree.heading(col, text=col.capitalize(), anchor="w", command=lambda c=col: toggle_sort(c))

    scrollb = ttk.Scrollbar(frame, orient = "vertical", command = tree.yview)
    scrollb.pack(side = "right", fill = "y")

    horiz_scroll = ttk.Scrollbar(frame, orient = "horizontal", command = tree.xview)
    horiz_scroll.pack(side = "bottom",fill = "x")

    tree.configure(yscrollcommand = scrollb.set)
    tree.configure(xscrollcommand = horiz_scroll.set)

    # Insert data into the Treeview
    for i, row in enumerate(df_prev.iterrows()):
        if i >= 1000:  # Limit to the first 1000 rows
            break
        tree.insert("", "end", values=list(row[1]))  # Adjusted to correctly unpack row data

    tree.pack(expand=True, fill="both")

    # Button to close the preview window
    close_b = tk.Button(prev_window, text="Close Preview", command=prev_window.destroy)
    close_b.pack(side="bottom", pady=10)
    return

# joining files together and converting func
def file_merge(data1,data2):
    # pandas take the already converted files and put into df format 
    df1 = dd.from_pandas(pd.DataFrame(data1), npartitions=1)
    df2 = dd.from_pandas(pd.DataFrame(data2), npartitions=1)

    # let user select join type 
    j_types = ['inner','left','right','outer']
    j_type = simpledialog.askstring("Join Type","Join Type Options: Inner, Left, Outer, Right:")
    j_type = j_type.lower() 
    
    # incase user mistype
    if j_type not in j_types:
        messagebox.showwarning("Invalid Selection","Please select a valid join type.")
        return

    # ask user for join type
    j_col = simpledialog.askstring("Join Condition","Enter the primary column to join on:")

    # incase of user error when choosing col
    if j_col not in df1.columns or j_col not in df2.columns:
        messagebox.showwarning("Invalid Column Choice","The specified column does not exist in both datasets.")
        return

    # test 
    # print("DATAFRAME TEST",df1)

    # merge files
    merge_dfs = dd.merge(df1,df2,how=j_type,on=j_col) # check for multi col abilities for on

    # let user choose format and name post merge
    output_type = simpledialog.askstring("Output File Type","Choose Output File Type (CSV, JSON, XML, EXCEL, PARQUET):")
    name = filedialog.asksaveasfilename(defaultextension=f".{output_type.lower()}",
                                        filetypes=[(f"{output_type} File",f".{output_type.lower()}")])
    
    # check file type to compute the change from the merged filetype
    # cant use my write funcs with this merge df from dask, need to use dask funcs (unfortunately)
    if output_type.lower() == "csv":
        merge_dfs.compute().to_csv(name,index = False)
    elif output_type.lower() == "xml":
        merge_dfs.compute().to_xml(name,index = False)
    elif output_type.lower() == "json":
        merge_dfs.compute().to_json(name,orient = 'records',lines = True) # same to json as my own func result
    elif output_type.lower() == "excel":
        merge_dfs.compute().to_excel(name,index = False)
    elif output_type.lower() == "parquet":
        merge_dfs.compute().to_parquet(name,index=False)
    else:
        warn = messagebox.askokcancel("No Valid File Format", icon="warning")
    return name

def create_plot(data, x_column, y_column=None, plot_type='bar'):

    """
    Create basic plots based on user input.

    Parameters:
    - data (pd.DataFrame): The dataset containing the data.
    - x_column (str): The column name to be used on the x-axis.
    - y_column (str, optional): The column name to be used on the y-axis for scatter and violin plots. 
                                Not used for bar plots.
    - plot_type (str): The type of plot to create. Options are 'bar', 'scatter', 'violin'.
    """

    if plot_type == 'bar':
        # For bar plots, we'll assume it's a count plot of categories if y_column is None
        plt.figure(figsize=(10, 6))
        sns.countplot(x=x_column, data=data)
        plt.title(f'Bar Plot of {x_column}')
    elif plot_type == 'scatter' and y_column is not None:
        # Scatter plot requires both x and y columns
        plt.figure(figsize=(10, 6))
        sns.scatterplot(x=x_column, y=y_column, data=data)
        plt.title(f'Scatter Plot of {x_column} vs {y_column}')
    elif plot_type == 'violin' and y_column is not None:
        # Violin plot requires both x and y columns
        plt.figure(figsize=(10, 6))
        sns.violinplot(x=x_column, y=y_column, data=data)
        plt.title(f'Violin Plot of {x_column} vs {y_column}')
    else:
        print("Invalid plot type or missing arguments.")
        return

    plt.xlabel(x_column)
    if y_column:
        plt.ylabel(y_column)
    plt.show()


def commands_p():
    # cli commands for terminal use of application
    pass

def winscp_conn():
    # allow users to interact with winscp or linux based system 
    # view and download files from linux server
    pass
